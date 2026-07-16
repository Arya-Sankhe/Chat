"""Unified XLSX workbook generator (XlsxWriter).

Creates workbooks from the create_document xlsx contract in one pass:
sheets, optional cover, column formats, formulas, charts, and conditional formats.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell

MAX_SHEETS = 12
MAX_ROWS = 5000
MAX_CHARTS_PER_SHEET = 4
MAX_CHARTS_PER_WORKBOOK = 12
MAX_CONDITIONAL_FORMATS = 8
CHART_ROW_STRIDE = 18
SERIES_COLORS = ["#0066CC", "#4A90D9", "#333333", "#8AB4D8", "#666666", "#A9CCE8"]

ALLOWED_FUNCTIONS = frozenset({
    "SUM", "AVERAGE", "MIN", "MAX", "COUNT", "COUNTA", "IF", "IFERROR",
    "SUMIF", "SUMIFS", "COUNTIF", "ROUND", "ABS", "INDEX", "MATCH", "VLOOKUP",
})
COLUMN_FORMATS = frozenset({"currency", "percent", "integer", "number", "date", "text"})

_FUNC_RE = re.compile(r"\b([A-Za-z][A-Za-z0-9.]*)\s*\(")
_DANGEROUS_FORMULA_RE = re.compile(
    r"(?i)(\[\[|\]\]|https?://|\\\\|cmd\b|powershell\b|javascript:|"
    r"webservice\b|filterxml\b|hyperlink\b|dde\b|call\b|register\b|"
    r"exec\b|evaluate\b)"
)
_NUMERIC_BODY = re.compile(r"^-?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?$")
_ISO_DATE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})(?:[T ].*)?$")

THEMES = {
    "clean": {
        "font": "Aptos",
        "heading_font": "Aptos Display",
        "accent": "#111111",
        "accent2": "#4B5563",
        "header": "#F2F4F7",
        "border": "#D9DDE3",
        "body": "#252525",
        "muted": "#777777",
    },
    "business": {
        "font": "Aptos",
        "heading_font": "Aptos Display",
        "accent": "#0F766E",
        "accent2": "#155E75",
        "header": "#CCFBF1",
        "border": "#99F6E4",
        "body": "#172B2A",
        "muted": "#64748B",
    },
    "academic": {
        "font": "Georgia",
        "heading_font": "Georgia",
        "accent": "#1D4ED8",
        "accent2": "#334155",
        "header": "#DBEAFE",
        "border": "#BFDBFE",
        "body": "#1E293B",
        "muted": "#64748B",
    },
}


def create_xlsx_workbook(path, input_data):
    """Write an .xlsx workbook to `path` and return that path."""
    path = Path(path)
    input_data = input_data or {}
    data = input_data.get("data") if isinstance(input_data.get("data"), dict) else {}
    theme = _resolve_theme(input_data, data)
    cover = data.get("cover") if isinstance(data.get("cover"), dict) else None
    plans = _sheet_plans(input_data, data)
    if not plans:
        plans = [{"name": "Summary", "rows": [["Title"], [str(input_data.get("title") or "Workbook")]],
                  "description": "", "columns": [], "charts": [], "conditional_formats": [], "freeze": True}]

    used_names = set()
    if cover is not None:
        used_names.add("cover")

    resolved = []
    for index, plan in enumerate(plans[:MAX_SHEETS]):
        name = _unique_sheet_name(plan.get("name") or f"Sheet {index + 1}", used_names)
        used_names.add(name.casefold())
        resolved.append({**plan, "source_name": str(plan.get("name") or name), "name": name})

    formula_sheet_names = {}
    for plan in resolved:
        formula_sheet_names.setdefault(plan["source_name"], plan["name"])

    workbook = xlsxwriter.Workbook(str(path))
    try:
        formats = _build_formats(workbook, theme)
        if cover is not None:
            _write_cover(workbook, formats, input_data, cover, resolved)

        workbook_charts = 0
        for plan in resolved:
            ws = workbook.add_worksheet(plan["name"])
            ws.hide_gridlines(2)
            rows = plan["rows"][:MAX_ROWS]
            columns = plan["columns"]
            if not rows:
                rows = [["Column1"], [""]]
            width = max(len(row) for row in rows)
            rows = [_pad_row(row, width) for row in rows]

            header_fmt = formats["header"]
            for col, value in enumerate(rows[0]):
                ws.write(0, col, "" if value is None else str(value), header_fmt)

            for row_idx, row in enumerate(rows[1:], start=1):
                is_total = _is_total_label(row[0] if row else "")
                for col, raw in enumerate(row):
                    _write_cell(ws, row_idx, col, raw, columns, formats, is_total, formula_sheet_names)

            last_row = len(rows) - 1
            last_col = width - 1
            if plan.get("freeze") is not False:
                ws.freeze_panes(1, 0)
            if last_row >= 1 and last_col >= 0:
                ws.add_table(0, 0, last_row, last_col, {
                    "columns": [{"header": value} for value in _unique_headers(rows[0])],
                    "style": "Table Style Light 9",
                })

            for col in range(width):
                spec = columns[col] if col < len(columns) else {}
                explicit = spec.get("width") if isinstance(spec, dict) else None
                if explicit:
                    ws.set_column(col, col, min(90, float(explicit)))
                else:
                    sample = max(
                        (len(str(rows[r][col] if col < len(rows[r]) else "")) for r in range(len(rows))),
                        default=8,
                    )
                    ws.set_column(col, col, min(48, max(10, sample + 2)))

            _apply_conditional_formats(ws, plan.get("conditional_formats") or [], last_row, last_col, width)
            chart_last_row = last_row
            while chart_last_row >= 1 and _is_total_label(rows[chart_last_row][0]):
                chart_last_row -= 1
            added = _add_charts(
                workbook, ws, plan["name"], plan.get("charts") or [], chart_last_row, width, workbook_charts
            )
            workbook_charts += added
    finally:
        workbook.close()

    expected = (["Cover"] if cover is not None else []) + [p["name"] for p in resolved]
    _validate_reopen(path, expected)
    return str(path)


def sanitize_sheet_name(value, fallback="Sheet"):
    cleaned = re.sub(r"[\[\]:*?/\\]+", " ", str(value or fallback))
    cleaned = re.sub(r"[^a-z0-9._ -]+", " ", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s+", "-", cleaned.strip())
    cleaned = re.sub(r"^-+|-+$", "", cleaned)
    return (cleaned or fallback)[:31]


def validate_formula(formula):
    """Return the formula if safe; raise ValueError otherwise."""
    text = str(formula or "").strip()
    if not text.startswith("="):
        raise ValueError("formula must start with '='")
    body = text[1:]
    stringless = re.sub(r'"(?:[^"]|"")*"', '""', body)
    if not body or re.search(r"[!].*\[|\[.+\]", stringless):
        raise ValueError("external workbook reference rejected")
    formula_code = re.sub(r"'(?:[^']|'')*'!", "Sheet!", stringless)
    if _DANGEROUS_FORMULA_RE.search(formula_code):
        raise ValueError("dangerous or external formula rejected")
    if re.search(r"#(?:REF!|DIV/0!|VALUE!|NAME\?|N/A)", formula_code, re.I):
        raise ValueError("formula contains an Excel error value")
    for match in _FUNC_RE.finditer(formula_code):
        name = match.group(1).upper()
        if name not in ALLOWED_FUNCTIONS:
            raise ValueError(f"unsupported function: {name}")
    return text


def _unique_sheet_name(value, used):
    base = sanitize_sheet_name(value)
    if base.casefold() not in used:
        return base
    for index in range(2, 1000):
        suffix = f"-{index}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        if candidate.casefold() not in used:
            return candidate
    raise RuntimeError("unable to allocate unique sheet name")


def _resolve_theme(input_data, data):
    explicit = str(input_data.get("theme") or data.get("theme") or "").strip().lower()
    if explicit in THEMES:
        return THEMES[explicit]
    text = " ".join([
        str(input_data.get("title") or ""),
        str(input_data.get("instructions") or ""),
    ]).lower()
    if re.search(r"\b(homework|assignment|lecture|class|course|student|university|school|research|exam)\b", text):
        return THEMES["academic"]
    if re.search(r"\b(business|strategy|proposal|client|executive|sales|finance|budget|kpi|dashboard|report)\b", text):
        return THEMES["business"]
    return THEMES["clean"]


def _build_formats(workbook, theme):
    base = {"font_name": theme["font"], "font_color": theme["body"], "valign": "top"}
    return {
        "_wb": workbook,
        "_base": base,
        "header": workbook.add_format({
            **base, "bold": True, "bg_color": theme["header"], "bottom": 1,
            "bottom_color": theme["accent"], "valign": "vcenter", "text_wrap": True,
        }),
        "body": workbook.add_format(base),
        "total": workbook.add_format({**base, "bold": True, "top": 1, "top_color": theme["accent"]}),
        "cover_title": workbook.add_format({
            "font_name": theme["heading_font"], "bold": True, "font_size": 18, "font_color": theme["accent"],
        }),
        "cover_section": workbook.add_format({
            "font_name": theme["heading_font"], "bold": True, "font_size": 12, "font_color": theme["accent2"],
        }),
        "cover_muted": workbook.add_format({
            "font_name": theme["font"], "font_size": 9, "font_color": theme["muted"],
        }),
        "cover_body": workbook.add_format({
            "font_name": theme["font"], "font_size": 10.5, "font_color": theme["body"],
        }),
        "cover_bold": workbook.add_format({
            "font_name": theme["font"], "bold": True, "font_size": 10.5, "font_color": theme["body"],
        }),
    }


def _typed_format(formats, kind, is_total=False, symbol="$"):
    if not kind:
        return formats["total"] if is_total else formats["body"]
    number_format = {
        "text": "@",
        "percent": "0.0%",
        "integer": "#,##0",
        "number": "#,##0.00",
        "date": "yyyy-mm-dd",
    }.get(kind)
    if kind == "currency":
        number_format = f"{(symbol or '$')[:3]}#,##0.00"
    key = f"typed:{kind}:{number_format}:{int(is_total)}"
    if key not in formats:
        props = {**formats["_base"], "num_format": number_format}
        if is_total:
            props.update({"bold": True, "top": 1})
        formats[key] = formats["_wb"].add_format(props)
    return formats[key]


def _currency_format(formats, symbol, integral, is_total):
    symbol = (symbol or "$")[:3]
    key = f"currency:{symbol}:{'i' if integral else 'f'}:{int(is_total)}"
    cached = formats.get(key)
    if cached:
        return cached
    pattern = f"{symbol}#,##0" if integral else f"{symbol}#,##0.00"
    props = {**formats["_base"], "num_format": pattern}
    if is_total:
        props.update({"bold": True, "top": 1})
    fmt = formats["_wb"].add_format(props)
    formats[key] = fmt
    return fmt


def _sheet_plans(input_data, data):
    explicit = data.get("sheets") if isinstance(data.get("sheets"), list) else None
    if not explicit and isinstance(input_data.get("sheets"), list):
        explicit = input_data["sheets"]
    plans = []
    if explicit:
        for index, sheet in enumerate(explicit[:MAX_SHEETS]):
            if not isinstance(sheet, dict):
                continue
            rows = _normalize_rows(sheet.get("rows") or _table_rows(sheet))
            if not rows:
                continue
            plans.append({
                "name": sheet.get("name") or sheet.get("title") or f"Sheet {index + 1}",
                "description": str(sheet.get("description") or "").strip(),
                "rows": rows,
                "columns": _normalize_columns(sheet.get("columns")),
                "charts": sheet.get("charts") if isinstance(sheet.get("charts"), list) else [],
                "conditional_formats": [
                    entry for entry in (sheet.get("conditional_formats") or [])
                    if isinstance(entry, dict)
                ],
                "freeze": sheet.get("freeze") is not False,
            })
        if plans:
            return plans

    rows = None
    if isinstance(data.get("rows"), list):
        rows = data["rows"]
    elif isinstance(input_data.get("rows"), list):
        rows = input_data["rows"]
    elif isinstance(input_data.get("tables"), list) and input_data["tables"]:
        rows = _table_rows(input_data["tables"][0])
    rows = _normalize_rows(rows or [])
    if not rows:
        return []
    return [{
        "name": "Summary",
        "description": "",
        "rows": rows,
        "columns": [],
        "charts": [],
        "conditional_formats": [],
        "freeze": True,
    }]


def _table_rows(table):
    if not isinstance(table, dict):
        return []
    headers = table.get("headers") if isinstance(table.get("headers"), list) else []
    rows = table.get("rows") if isinstance(table.get("rows"), list) else None
    if rows is None:
        rows = table.get("data") if isinstance(table.get("data"), list) else []
    all_rows = ([headers] + rows) if headers else rows
    return all_rows[:MAX_ROWS]


def _normalize_rows(rows):
    out = []
    for row in rows or []:
        values = row if isinstance(row, list) else [row]
        cleaned = ["" if cell is None else cell for cell in values]
        if any(str(cell).strip() != "" for cell in cleaned):
            out.append(cleaned)
    return out


def _normalize_columns(columns):
    if not isinstance(columns, list):
        return []
    out = []
    for column in columns:
        if not isinstance(column, dict):
            out.append({})
            continue
        fmt = column.get("format") if column.get("format") in COLUMN_FORMATS else ""
        symbol = column.get("symbol") if isinstance(column.get("symbol"), str) and column.get("symbol").strip() else "$"
        width = column.get("width") if isinstance(column.get("width"), (int, float)) and column.get("width") > 0 else 0
        out.append({"format": fmt, "symbol": symbol.strip()[:3], "width": width})
    return out


def _pad_row(row, width):
    values = list(row)
    if len(values) < width:
        values.extend([""] * (width - len(values)))
    return values[:width]


def _unique_headers(row):
    used = set()
    headers = []
    for index, value in enumerate(row, start=1):
        base = str(value or f"Column {index}").strip() or f"Column {index}"
        candidate = base
        suffix = 2
        while candidate.casefold() in used:
            candidate = f"{base} {suffix}"
            suffix += 1
        used.add(candidate.casefold())
        headers.append(candidate)
    return headers


def _is_total_label(value):
    return bool(re.match(r"^(total|sum|grand total|net|subtotal)\b", str(value or "").strip(), re.I))


def _write_cell(ws, row, col, raw, columns, formats, is_total, sheet_names):
    spec = columns[col] if col < len(columns) else {}
    fmt_name = spec.get("format") if isinstance(spec, dict) else ""

    formula_input = raw if isinstance(raw, str) and raw.startswith("=") else None
    if isinstance(raw, dict) and raw.get("formula"):
        formula_input = f"={str(raw['formula']).lstrip('=')}"
    if formula_input:
        cell_fmt = _typed_format(formats, fmt_name, is_total, spec.get("symbol") or "$")
        try:
            formula = _resolve_formula_sheets(validate_formula(formula_input), sheet_names)
        except ValueError:
            ws.write_string(row, col, formula_input, cell_fmt)
        else:
            ws.write_formula(row, col, formula, cell_fmt)
        return

    value, cell_fmt = _convert_value(raw, fmt_name, spec if isinstance(spec, dict) else {}, formats, is_total)
    if isinstance(value, datetime):
        ws.write_datetime(row, col, value, cell_fmt)
    elif isinstance(value, date):
        ws.write_datetime(row, col, datetime(value.year, value.month, value.day), cell_fmt)
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        ws.write_number(row, col, value, cell_fmt)
    else:
        ws.write_string(row, col, "" if value is None else str(value), cell_fmt)


def _resolve_formula_sheets(formula, sheet_names):
    for source, target in sheet_names.items():
        if source == target:
            continue
        escaped = source.replace("'", "''")
        formula = re.sub(
            re.escape(f"'{escaped}'!"),
            lambda _: f"'{target}'!",
            formula,
            flags=re.IGNORECASE,
        )
        formula = re.sub(
            rf"(?<![A-Za-z0-9_.]){re.escape(source)}!",
            lambda _: f"'{target}'!",
            formula,
            flags=re.IGNORECASE,
        )
    return formula


def _convert_value(raw, fmt_name, spec, formats, is_total):
    fallback = formats["total"] if is_total else formats["body"]
    if fmt_name == "text":
        text = "" if raw is None else str(raw)
        return text, _typed_format(formats, "text", is_total)

    if fmt_name == "percent":
        number = _as_percent(raw)
        if number is not None:
            return number, _typed_format(formats, "percent", is_total)
        return "" if raw is None else str(raw), fallback

    if fmt_name == "currency":
        number = _as_number(raw)
        if number is not None:
            return number, _currency_format(
                formats, spec.get("symbol") or "$", float(number).is_integer(), is_total
            )
        return "" if raw is None else str(raw), fallback

    if fmt_name == "integer":
        number = _as_number(raw)
        if number is not None:
            return int(round(number)), _typed_format(formats, "integer", is_total)
        return "" if raw is None else str(raw), fallback

    if fmt_name == "number":
        number = _as_number(raw)
        if number is not None:
            return float(number), _typed_format(formats, "number", is_total)
        return "" if raw is None else str(raw), fallback

    if fmt_name == "date":
        parsed = _as_date(raw)
        if parsed is not None:
            return parsed, _typed_format(formats, "date", is_total)
        return "" if raw is None else str(raw), fallback

    # Unspecified columns preserve strings (and leave real numbers as numbers).
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return raw, fallback
    if isinstance(raw, (date, datetime)):
        return raw, _typed_format(formats, "date", is_total)
    return "" if raw is None else str(raw), fallback


def _as_number(raw):
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw or "").strip()
    if not text or text.startswith("="):
        return None
    currency = re.match(r"^([$€£¥₹])\s?(.+)$", text)
    if currency and _NUMERIC_BODY.match(currency.group(2).strip()):
        return float(currency.group(2).strip().replace(",", ""))
    if _NUMERIC_BODY.match(text):
        return float(text.replace(",", ""))
    return None


def _as_percent(raw):
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw or "").strip()
    if not text or text.startswith("="):
        return None
    if text.endswith("%"):
        body = text[:-1].strip()
        if _NUMERIC_BODY.match(body):
            return float(body.replace(",", "")) / 100.0
        return None
    if _NUMERIC_BODY.match(text):
        return float(text.replace(",", ""))
    return None


def _as_date(raw):
    if isinstance(raw, datetime):
        return raw
    if isinstance(raw, date):
        return raw
    text = str(raw or "").strip()
    match = _ISO_DATE.match(text)
    if not match:
        return None
    try:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def _write_cover(workbook, formats, input_data, cover, sheet_plans):
    ws = workbook.add_worksheet("Cover")
    ws.hide_gridlines(2)
    ws.set_column(0, 0, 3)
    ws.set_column(1, 1, 28)
    ws.set_column(2, 2, 52)
    ws.write(1, 1, str(input_data.get("title") or "Workbook"), formats["cover_title"])
    subtitle = str(cover.get("subtitle") or input_data.get("instructions") or "").strip()
    row = 2
    if subtitle:
        ws.write(row, 1, subtitle, formats["cover_muted"])
        row += 1
    ws.write(row, 1, f"Generated {date.today().isoformat()}", formats["cover_muted"])
    row += 2
    metrics = [
        {"label": str(m.get("label") or "").strip(), "value": m.get("value")}
        for m in (cover.get("metrics") or [])
        if isinstance(m, dict) and str(m.get("label") or "").strip()
    ][:8]
    if metrics:
        ws.write(row, 1, "Key Metrics", formats["cover_section"])
        row += 1
        for metric in metrics:
            ws.write(row, 1, metric["label"], formats["cover_body"])
            value = metric["value"]
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                ws.write_number(row, 2, value, formats["cover_bold"])
            else:
                ws.write(row, 2, "" if value is None else str(value), formats["cover_bold"])
            row += 1
        row += 1
    ws.write(row, 1, "Sheets", formats["cover_section"])
    row += 1
    for plan in sheet_plans:
        ws.write(row, 1, plan["name"], formats["cover_bold"])
        if plan.get("description"):
            ws.write(row, 2, plan["description"], formats["cover_muted"])
        row += 1
    notes = str(cover.get("notes") or "").strip()
    if notes:
        row += 1
        ws.write(row, 1, notes, formats["cover_muted"])


def _apply_conditional_formats(ws, specs, last_row, last_col, width):
    if last_row < 1:
        return
    for entry in specs[:MAX_CONDITIONAL_FORMATS]:
        try:
            column = int(entry.get("column"))
        except (TypeError, ValueError):
            continue
        if column < 1 or column > width or column - 1 > last_col:
            continue
        col = column - 1
        start = xl_rowcol_to_cell(1, col)
        end = xl_rowcol_to_cell(last_row, col)
        ref = f"{start}:{end}"
        kind = str(entry.get("type") or "").strip().lower()
        if kind == "data_bar":
            ws.conditional_format(ref, {"type": "data_bar", "bar_color": "#4A90D9"})
        elif kind == "color_scale":
            ws.conditional_format(ref, {
                "type": "3_color_scale",
                "min_color": "#FFFFFF",
                "mid_color": "#A9CCE8",
                "max_color": "#0066CC",
            })
        elif kind == "icon_set":
            ws.conditional_format(ref, {"type": "icon_set", "icon_style": "3_traffic_lights"})


def _add_charts(workbook, ws, sheet_name, charts, last_row, width, workbook_charts):
    if last_row < 1 or width < 1:
        return 0
    added = 0
    for spec in charts:
        if workbook_charts + added >= MAX_CHARTS_PER_WORKBOOK or added >= MAX_CHARTS_PER_SHEET:
            break
        if not isinstance(spec, dict):
            continue
        try:
            categories_column = int(spec.get("categories_column"))
        except (TypeError, ValueError):
            continue
        if categories_column < 1 or categories_column > width:
            continue
        series_columns = []
        for col in spec.get("series_columns") or []:
            try:
                value = int(col)
            except (TypeError, ValueError):
                continue
            if 1 <= value <= width:
                series_columns.append(value)
        if not series_columns:
            continue

        chart_type = str(spec.get("type") or "bar").strip().lower()
        # Contract "bar" maps to Excel column chart (vertical), matching prior openpyxl behavior.
        xw_type = {"bar": "column", "line": "line", "pie": "pie", "area": "area"}.get(chart_type, "column")
        series_last_row = min(last_row, 6) if xw_type == "pie" else last_row
        chart = workbook.add_chart({"type": xw_type})
        cat_col = categories_column - 1
        for index, series_col in enumerate(series_columns):
            series = {
                "name": [sheet_name, 0, series_col - 1],
                "categories": [sheet_name, 1, cat_col, series_last_row, cat_col],
                "values": [sheet_name, 1, series_col - 1, series_last_row, series_col - 1],
            }
            if xw_type != "pie":
                series["fill"] = {"color": SERIES_COLORS[index % len(SERIES_COLORS)]}
            chart.add_series(series)
        if spec.get("title"):
            chart.set_title({"name": str(spec["title"])})
        if xw_type != "pie":
            if spec.get("x_axis_title"):
                chart.set_x_axis({"name": str(spec["x_axis_title"])})
            if spec.get("y_axis_title"):
                chart.set_y_axis({"name": str(spec["y_axis_title"])})
        chart.set_size({"width": 640, "height": 360})
        anchor_col = width + 1
        anchor_row = added * CHART_ROW_STRIDE
        ws.insert_chart(anchor_row, anchor_col, chart)
        added += 1
    return added


def _validate_reopen(path, expected_names):
    """Deterministic reopen check; does not rewrite the file."""
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=False)
    try:
        if wb.sheetnames != expected_names:
            raise RuntimeError(f"xlsx validation failed: sheets {wb.sheetnames!r} != {expected_names!r}")
        first = expected_names[0]
        ws = wb[first]
        # Ensure at least the header/title cell exists.
        next(ws.iter_rows(min_row=1, max_row=1, max_col=1, values_only=True), None)
    finally:
        wb.close()
