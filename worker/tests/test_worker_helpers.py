import json
import os
import shutil
import tempfile
import threading
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook

from worker import worker as w


class EnvHelpersTest(unittest.TestCase):
    def test_env_int_clamps_to_bounds(self):
        with mock.patch.dict(os.environ, {"TEST_INT": "99"}, clear=False):
            self.assertEqual(w.env_int("TEST_INT", 2, minimum=1, maximum=4), 4)
        with mock.patch.dict(os.environ, {"TEST_INT": "0"}, clear=False):
            self.assertEqual(w.env_int("TEST_INT", 2, minimum=1, maximum=4), 1)
        with mock.patch.dict(os.environ, {"TEST_INT": "nope"}, clear=False):
            self.assertEqual(w.env_int("TEST_INT", 2, minimum=1, maximum=4), 2)

    def test_default_lease_heartbeat_seconds(self):
        self.assertEqual(w.default_lease_heartbeat_seconds(120), 30.0)
        self.assertEqual(w.default_lease_heartbeat_seconds(8), 5.0)
        self.assertEqual(w.default_lease_heartbeat_seconds(200), 30.0)

    def test_worker_concurrency_default_and_cap(self):
        with mock.patch.dict(os.environ, {}, clear=False):
            os.environ.pop("DOCUMENT_WORKER_CONCURRENCY", None)
            self.assertEqual(w.worker_concurrency(), 1)
        with mock.patch.dict(os.environ, {"DOCUMENT_WORKER_CONCURRENCY": "100"}, clear=False):
            self.assertEqual(w.worker_concurrency(), w.WORKER_CONCURRENCY_CAP)


class RetryHelpersTest(unittest.TestCase):
    def test_is_retryable_http_status(self):
        self.assertTrue(w.is_retryable_http_status(429))
        self.assertTrue(w.is_retryable_http_status(500))
        self.assertTrue(w.is_retryable_http_status(503))
        self.assertFalse(w.is_retryable_http_status(400))
        self.assertFalse(w.is_retryable_http_status(404))
        self.assertFalse(w.is_retryable_http_status(200))

    def test_request_with_retries_retries_429_then_succeeds(self):
        responses = [
            mock.Mock(ok=False, status_code=429, text="slow down"),
            mock.Mock(ok=True, status_code=200, text="{}"),
        ]
        with mock.patch("worker.worker.requests.request", side_effect=responses) as request_mock:
            with mock.patch("worker.worker.time.sleep") as sleep_mock:
                result = w.request_with_retries("GET", "https://example.test", max_attempts=3)
        self.assertIs(result, responses[1])
        self.assertEqual(request_mock.call_count, 2)
        sleep_mock.assert_called_once()

    def test_request_with_retries_does_not_retry_permanent_4xx(self):
        response = mock.Mock(ok=False, status_code=400, text="bad")
        with mock.patch("worker.worker.requests.request", return_value=response) as request_mock:
            with mock.patch("worker.worker.time.sleep") as sleep_mock:
                result = w.request_with_retries("GET", "https://example.test", max_attempts=4)
        self.assertIs(result, response)
        self.assertEqual(request_mock.call_count, 1)
        sleep_mock.assert_not_called()

    def test_request_with_retries_retries_network_errors(self):
        import requests

        responses = [
            requests.exceptions.ConnectionError("boom"),
            mock.Mock(ok=True, status_code=200, text="{}"),
        ]
        with mock.patch("worker.worker.requests.request", side_effect=responses) as request_mock:
            with mock.patch("worker.worker.time.sleep"):
                result = w.request_with_retries("GET", "https://example.test", max_attempts=3)
        self.assertTrue(result.ok)
        self.assertEqual(request_mock.call_count, 2)


class XlsxRecalculationTest(unittest.TestCase):
    def test_create_xlsx_recalculates_before_delivery(self):
        processor = w.Processor.__new__(w.Processor)
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            recalculated = tmp_path / "recalculated" / "Budget.xlsx"
            processor.recalculate_xlsx = mock.Mock(return_value=recalculated)
            result = processor.create_xlsx(tmp_path, "Budget", {
                "data": {"sheets": [{"name": "Costs", "rows": [["Value"], [1]]}]}
            })
        self.assertEqual(result, recalculated)
        processor.recalculate_xlsx.assert_called_once()

    def test_recalculation_rejects_missing_formula_results(self):
        processor = w.Processor.__new__(w.Processor)
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "formula.xlsx"
            workbook = Workbook()
            workbook.active["A1"] = "=1+1"
            workbook.save(source)

            def copy_without_recalculation(command, **_kwargs):
                output = tmp_path / "recalculated" / source.name
                output.parent.mkdir(parents=True, exist_ok=True)
                shutil.copyfile(source, output)
                return mock.Mock(returncode=0)

            with mock.patch("worker.worker.subprocess.run", side_effect=copy_without_recalculation):
                with self.assertRaisesRegex(RuntimeError, "xlsx_formula_error"):
                    processor.recalculate_xlsx(source, tmp_path)


class SplitPageRangesTest(unittest.TestCase):
    def test_single_range_when_workers_not_justified(self):
        self.assertEqual(w.split_page_ranges(3, 2), [(1, 3)])
        self.assertEqual(w.split_page_ranges(10, 1), [(1, 10)])
        self.assertEqual(w.split_page_ranges(0, 2), [])

    def test_splits_evenly_when_justified(self):
        self.assertEqual(w.split_page_ranges(10, 2), [(1, 5), (6, 10)])
        self.assertEqual(w.split_page_ranges(11, 3), [(1, 4), (5, 8), (9, 11)])
        self.assertEqual(w.split_page_ranges(8, 4), [(1, 2), (3, 4), (5, 6), (7, 8)])


class EdgeParseNormalizeTest(unittest.TestCase):
    SAMPLE = {
        "file name": "demo.pdf",
        "number of pages": 2,
        "kids": [
            {
                "type": "heading",
                "heading level": 1,
                "page number": 1,
                "content": "Intro",
            },
            {
                "type": "paragraph",
                "page number": 1,
                "content": "Hello world.",
            },
            {
                "type": "table",
                "page number": 2,
                "rows": [
                    {
                        "row number": 1,
                        "cells": [
                            {"kids": [{"type": "paragraph", "content": "A"}]},
                            {"kids": [{"type": "paragraph", "content": "B"}]},
                        ],
                    },
                    {
                        "row number": 2,
                        "cells": [
                            {"content": "1"},
                            {"content": "2"},
                        ],
                    },
                ],
            },
        ],
    }

    def test_normalize_accepts_json_string(self):
        normalized = w.normalize_edgeparse_result(json.dumps(self.SAMPLE))
        self.assertEqual(normalized["number of pages"], 2)
        self.assertEqual(len(normalized["kids"]), 3)

    def test_normalize_accepts_object_wrapper(self):
        wrapper = mock.Mock()
        wrapper.json = self.SAMPLE
        for attr in ("data", "document", "result", "output"):
            setattr(wrapper, attr, None)
        normalized = w.normalize_edgeparse_result(wrapper)
        self.assertEqual(normalized["number of pages"], 2)
        self.assertEqual(normalized["kids"][0]["content"], "Intro")

    def test_group_pages_builds_structured_markdown(self):
        grouped = w.group_edgeparse_pages(self.SAMPLE)
        self.assertEqual(grouped["page_count"], 2)
        self.assertTrue(grouped["has_usable_text"])
        page1 = grouped["pages"][0]["text"]
        page2 = grouped["pages"][1]["text"]
        self.assertIn("# Intro", page1)
        self.assertIn("Hello world.", page1)
        self.assertIn("| A | B |", page2)
        self.assertIn("| 1 | 2 |", page2)

    def test_empty_text_semantics(self):
        empty = {
            "number of pages": 1,
            "kids": [
                {"type": "image", "page number": 1, "source": "x.png"},
                {"type": "paragraph", "page number": 1, "content": "   "},
            ],
        }
        grouped = w.group_edgeparse_pages(empty)
        self.assertFalse(grouped["has_usable_text"])
        self.assertEqual(grouped["word_count"], 0)
        self.assertEqual(grouped["pages"][0]["text"], "")

    def test_group_pages_trusts_max_page_count(self):
        inflated = {
            "number of pages": 99,
            "kids": [
                {"type": "paragraph", "page number": 1, "content": "Page one text here."},
                {"type": "paragraph", "page number": 50, "content": "Should be ignored."},
            ],
        }
        grouped = w.group_edgeparse_pages(inflated, max_page_count=2)
        self.assertEqual(grouped["page_count"], 2)
        self.assertEqual(len(grouped["pages"]), 2)
        self.assertIn("Page one text here.", grouped["pages"][0]["text"])
        self.assertEqual(grouped["pages"][1]["text"], "")

    def test_single_garbage_token_is_not_usable(self):
        self.assertFalse(w.is_usable_extracted_text("x"))
        self.assertFalse(w.is_usable_extracted_text("ab"))
        self.assertTrue(w.is_usable_extracted_text("one two three"))
        self.assertTrue(w.is_usable_extracted_text("abcdefghijklmnopqrst"))  # 20 alnum


class StageAwareCompletionPayloadTest(unittest.TestCase):
    def test_empty_extract_patch_omits_text_ready_and_records_warning(self):
        processor = w.Processor.__new__(w.Processor)
        processor.default_limits = {"max_pdf_pages": 100, "max_extracted_chars": 500000}
        processor.max_extracted_chars = 500000
        processor.object_key = lambda user_id, name: f"users/{user_id}/{name}"

        doc = {
            "id": "doc-1",
            "user_id": "user-1",
            "attachment_id": "att-1",
            "kind": "pdf",
        }
        attachment = {
            "id": "att-1",
            "file_name": "scan.pdf",
            "object_key": "raw/scan.pdf",
            "content_type": "application/pdf",
            "size_bytes": 10,
            "etag": "etag",
        }
        job = {"id": "job-1", "document_file_id": "doc-1", "input": {}}

        processor.db = mock.Mock()
        processor.db.update_document_file.return_value = doc
        processor.db.insert_chunks.return_value = None
        processor.r2 = mock.Mock()
        processor.r2.upload.return_value = "etag"
        processor.assert_job_active = mock.Mock()
        processor.extract_pdf = mock.Mock(return_value=([], {
            "page_count": 2,
            "word_count": 0,
            "has_usable_text": False,
            "extractor": "edgeparse",
        }))

        with mock.patch("worker.worker.PdfReader") as reader_cls:
            reader = reader_cls.return_value
            reader.is_encrypted = False
            reader.pages = [object(), object()]
            result = processor.extract_pdf_text_job(
                job,
                Path(tempfile.mkdtemp()),
                doc,
                attachment,
                Path("/tmp/scan.pdf"),
                processor.default_limits,
            )

        patch = result["_document_patch"]
        self.assertNotIn("text_ready_at", patch)
        self.assertEqual(patch["page_count"], 2)
        self.assertEqual(patch["word_count"], 0)
        self.assertIsNone(patch["error"])
        self.assertEqual(patch["stage_errors"]["text"]["code"], "no_usable_digital_text")
        self.assertEqual(result["status"], "text_incapable")
        processor.db.delete_chunks.assert_not_called()
        processor.db.insert_chunks.assert_called_once()

    def test_edgeparse_failure_is_soft_text_incapable(self):
        processor = w.Processor.__new__(w.Processor)
        processor.default_limits = {"max_pdf_pages": 100, "max_extracted_chars": 500000}
        processor.max_extracted_chars = 500000
        processor.object_key = lambda user_id, name: f"users/{user_id}/{name}"
        processor.db = mock.Mock()
        processor.r2 = mock.Mock()
        processor.r2.upload.return_value = "etag"
        processor.assert_job_active = mock.Mock()
        processor.extract_pdf = mock.Mock(return_value=([], {
            "page_count": 3,
            "word_count": 0,
            "has_usable_text": False,
            "extractor": "edgeparse",
            "edgeparse_failed": True,
            "edgeparse_error": "boom",
        }))

        with mock.patch("worker.worker.PdfReader") as reader_cls:
            reader = reader_cls.return_value
            reader.is_encrypted = False
            reader.pages = [object(), object(), object()]
            result = processor.extract_pdf_text_job(
                {"id": "job-ep", "input": {}},
                Path(tempfile.mkdtemp()),
                {"id": "doc-1", "user_id": "user-1", "kind": "pdf"},
                {
                    "file_name": "a.pdf",
                    "content_type": "application/pdf",
                    "size_bytes": 1,
                    "etag": "e",
                },
                Path("/tmp/a.pdf"),
                processor.default_limits,
            )

        patch = result["_document_patch"]
        self.assertEqual(result["status"], "text_incapable")
        self.assertNotIn("text_ready_at", patch)
        self.assertIsNone(patch["error"])
        self.assertEqual(patch["stage_errors"]["text"]["code"], "edgeparse_failed")
        processor.db.delete_chunks.assert_not_called()

    def test_extract_pdf_soft_fails_on_edgeparse_convert_error(self):
        processor = w.Processor.__new__(w.Processor)
        processor.default_limits = {"max_pdf_pages": 100, "max_extracted_chars": 500000}
        processor.max_extracted_chars = 500000
        processor.cap_chunks = w.Processor.cap_chunks.__get__(processor, w.Processor)
        processor.limit = w.Processor.limit.__get__(processor, w.Processor)
        processor.chunk = w.Processor.chunk.__get__(processor, w.Processor)

        with mock.patch("worker.worker.PdfReader") as reader_cls:
            reader = reader_cls.return_value
            reader.is_encrypted = False
            reader.pages = [object(), object()]
            with mock.patch("worker.worker.edgeparse.convert", side_effect=RuntimeError("parse boom")):
                chunks, meta = processor.extract_pdf(
                    "/tmp/a.pdf",
                    "user-1",
                    "doc-1",
                    processor.default_limits,
                    page_count=2,
                )

        self.assertEqual(chunks, [])
        self.assertTrue(meta["edgeparse_failed"])
        self.assertEqual(meta["page_count"], 2)
        self.assertFalse(meta["has_usable_text"])
        self.assertIn("parse boom", meta["edgeparse_error"])

    def test_extract_pdf_uses_documented_edgeparse_call(self):
        processor = w.Processor.__new__(w.Processor)
        processor.default_limits = {"max_pdf_pages": 100, "max_extracted_chars": 500000}
        processor.max_extracted_chars = 500000
        processor.cap_chunks = w.Processor.cap_chunks.__get__(processor, w.Processor)
        processor.limit = w.Processor.limit.__get__(processor, w.Processor)
        processor.chunk = w.Processor.chunk.__get__(processor, w.Processor)
        sample = {
            "number of pages": 1,
            "kids": [
                {"type": "paragraph", "page number": 1, "content": "Alpha beta gamma delta."},
            ],
        }

        with mock.patch("worker.worker.PdfReader") as reader_cls:
            reader = reader_cls.return_value
            reader.is_encrypted = False
            reader.pages = [object()]
            with mock.patch("worker.worker.edgeparse.convert", return_value=sample) as convert_mock:
                chunks, meta = processor.extract_pdf(
                    "/tmp/a.pdf",
                    "user-1",
                    "doc-1",
                    processor.default_limits,
                    page_count=1,
                )

        convert_mock.assert_called_once_with("/tmp/a.pdf", format="json")
        self.assertTrue(meta["has_usable_text"])
        self.assertEqual(meta["page_count"], 1)
        self.assertEqual(len(chunks), 1)

    def test_usable_extract_patch_sets_text_ready(self):
        processor = w.Processor.__new__(w.Processor)
        processor.default_limits = {"max_pdf_pages": 100, "max_extracted_chars": 500000}
        processor.max_extracted_chars = 500000
        processor.object_key = lambda user_id, name: f"users/{user_id}/{name}"
        processor.db = mock.Mock()
        processor.r2 = mock.Mock()
        processor.r2.upload.return_value = "etag"
        processor.assert_job_active = mock.Mock()
        processor.extract_pdf = mock.Mock(return_value=([{
            "user_id": "user-1",
            "document_file_id": "doc-1",
            "chunk_index": 0,
            "source_type": "page",
            "source_label": "Page 1",
            "text": "Hello there friend",
            "char_count": 17,
            "token_estimate": 4,
            "metadata": {"page": 1},
        }], {
            "page_count": 1,
            "word_count": 3,
            "has_usable_text": True,
            "extractor": "edgeparse",
        }))

        with mock.patch("worker.worker.PdfReader") as reader_cls:
            reader = reader_cls.return_value
            reader.is_encrypted = False
            reader.pages = [object()]
            result = processor.extract_pdf_text_job(
                {"id": "job-1", "input": {}},
                Path(tempfile.mkdtemp()),
                {"id": "doc-1", "user_id": "user-1", "kind": "pdf"},
                {
                    "file_name": "a.pdf",
                    "content_type": "application/pdf",
                    "size_bytes": 1,
                    "etag": "e",
                },
                Path("/tmp/a.pdf"),
                processor.default_limits,
            )

        self.assertIn("text_ready_at", result["_document_patch"])
        self.assertNotIn("stage_errors", result["_document_patch"])
        self.assertEqual(result["status"], "text_ready")
        processor.db.delete_chunks.assert_not_called()

    def _enrich_processor(self, *, embeddings_enabled=True, embed_side_effect=None, embed_return=None):
        processor = w.Processor.__new__(w.Processor)
        processor.default_limits = {"max_pdf_pages": 100}
        processor.visual_page_dpi = 144
        processor.page_upload_workers = 2
        processor.pdf_render_workers = 1
        processor.worker_id = "worker-a"
        processor.db = mock.Mock()
        processor.db.get_document_file.side_effect = lambda *_a, **_k: {
            "id": "doc-1",
            "user_id": "user-1",
            "attachment_id": "att-1",
            "kind": "pdf",
        }
        processor.db.get_attachment.return_value = {
            "id": "att-1",
            "file_name": "a.pdf",
            "object_key": "raw/a.pdf",
            "content_type": "application/pdf",
            "size_bytes": 1,
            "etag": "e",
        }
        processor.db.publish_document_visual_ready.return_value = {
            "id": "doc-1",
            "visual_ready_at": "2026-07-11T12:00:00+00:00",
        }
        processor.r2 = mock.Mock()
        processor.embeddings = mock.Mock()
        processor.embeddings.enabled = embeddings_enabled
        processor.embeddings.model = "jina-test"
        if embed_side_effect is not None:
            processor.embeddings.embed_images.side_effect = embed_side_effect
        elif embed_return is not None:
            processor.embeddings.embed_images.return_value = embed_return
        processor.assert_job_active = mock.Mock()
        processor.estimated_page_pixels = mock.Mock(return_value=(100, 200))
        return processor

    def test_enrich_patch_sets_visual_ready_without_enriched_on_jina_failure(self):
        processor = self._enrich_processor(embed_side_effect=RuntimeError("jina down"))

        with tempfile.TemporaryDirectory() as tmp:
            image = Path(tmp) / "page-1.jpg"
            image.write_bytes(b"jpg")
            processor.render_pdf_pages = mock.Mock(return_value=[image])
            with mock.patch("worker.worker.PdfReader") as reader_cls:
                reader = reader_cls.return_value
                reader.is_encrypted = False
                page = mock.Mock()
                reader.pages = [page]
                result = processor.enrich_pdf_job(
                    {"id": "job-2", "document_file_id": "doc-1", "input": {}},
                    Path(tmp),
                )

        patch = result["_document_patch"]
        self.assertIn("visual_ready_at", patch)
        self.assertNotIn("enriched_at", patch)
        self.assertEqual(patch["stage_errors"]["enrichment"]["code"], "embedding_failed")
        self.assertTrue(result["status"] == "visual_ready")
        self.assertFalse(result["enriched"])
        processor.db.publish_document_visual_ready.assert_called_once()
        self.assertEqual(processor.db.publish_document_visual_ready.call_args.args[0], "job-2")
        self.assertEqual(processor.db.publish_document_visual_ready.call_args.args[3], 1)
        processor.embeddings.embed_images.assert_called_once()

    def test_office_enrich_uses_temporary_pdf_without_persisting_it(self):
        processor = self._enrich_processor(embeddings_enabled=False)
        processor.db.get_document_file.side_effect = lambda *_a, **_k: {
            "id": "doc-1",
            "user_id": "user-1",
            "attachment_id": "att-1",
            "kind": "pptx",
        }
        processor.db.get_attachment.return_value = {
            "id": "att-1",
            "file_name": "deck.pptx",
            "object_key": "raw/deck.pptx",
            "content_type": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
            "size_bytes": 1,
            "etag": "e",
        }

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            temporary_pdf = tmp_path / "out" / "deck.pdf"
            temporary_pdf.parent.mkdir()
            temporary_pdf.write_bytes(b"pdf")
            image = tmp_path / "page-1.jpg"
            image.write_bytes(b"jpg")
            processor.libreoffice_convert = mock.Mock(return_value=temporary_pdf)
            processor.render_pdf_pages = mock.Mock(return_value=[image])
            with mock.patch("worker.worker.PdfReader") as reader_cls:
                reader = reader_cls.return_value
                reader.is_encrypted = False
                reader.pages = [mock.Mock()]
                result = processor.enrich_pdf_job(
                    {"id": "job-2", "document_file_id": "doc-1", "input": {}},
                    tmp_path,
                )

        processor.libreoffice_convert.assert_called_once_with(tmp_path / "deck.pptx", tmp_path, "pdf")
        self.assertEqual(processor.render_pdf_pages.call_args.args[0], temporary_pdf)
        uploaded_keys = [call.args[0] for call in processor.r2.upload.call_args_list]
        self.assertEqual(uploaded_keys, ["users/user-1/documents/doc-1/pages/page-0001.jpg"])
        self.assertEqual(result["_document_patch"]["metadata"]["source_kind"], "pptx")

    def test_pdf_enrich_does_not_call_libreoffice(self):
        processor = self._enrich_processor(embeddings_enabled=False)
        processor.libreoffice_convert = mock.Mock()

        with tempfile.TemporaryDirectory() as tmp:
            image = Path(tmp) / "page-1.jpg"
            image.write_bytes(b"jpg")
            processor.render_pdf_pages = mock.Mock(return_value=[image])
            with mock.patch("worker.worker.PdfReader") as reader_cls:
                reader = reader_cls.return_value
                reader.is_encrypted = False
                reader.pages = [mock.Mock()]
                processor.enrich_pdf_job(
                    {"id": "job-2", "document_file_id": "doc-1", "input": {}},
                    Path(tmp),
                )

        processor.libreoffice_convert.assert_not_called()

    def test_enrich_publishes_visual_ready_before_jina(self):
        order = []
        processor = self._enrich_processor(embed_return=[[0.1] * 8])

        def track_publish(*args, **kwargs):
            order.append("publish")
            return {"id": "doc-1", "visual_ready_at": "2026-07-11T12:00:00+00:00"}

        def track_embed(paths):
            order.append("embed")
            return [[0.1] * 8 for _ in paths]

        processor.db.publish_document_visual_ready.side_effect = track_publish
        processor.embeddings.embed_images.side_effect = track_embed

        with tempfile.TemporaryDirectory() as tmp:
            image = Path(tmp) / "page-1.jpg"
            image.write_bytes(b"jpg")
            processor.render_pdf_pages = mock.Mock(return_value=[image])
            with mock.patch("worker.worker.PdfReader") as reader_cls:
                reader = reader_cls.return_value
                reader.is_encrypted = False
                reader.pages = [mock.Mock()]
                result = processor.enrich_pdf_job(
                    {"id": "job-2", "document_file_id": "doc-1", "input": {}},
                    Path(tmp),
                )

        self.assertEqual(order, ["publish", "embed"])
        self.assertIn("enriched_at", result["_document_patch"])
        self.assertTrue(result["enriched"])

    def test_enrich_publishes_finished_render_ranges_without_waiting_for_all_ranges(self):
        processor = self._enrich_processor(embeddings_enabled=False)
        processor.pdf_render_workers = 2
        second_range_release = threading.Event()
        rendered_ranges = []

        def render_range(_source, output_dir, _dpi, **kwargs):
            first = kwargs["first_page"]
            last = kwargs["last_page"]
            if first == 3:
                self.assertTrue(second_range_release.wait(2), "first range should upload before the second finishes")
            paths = []
            for page_number in range(first, last + 1):
                image = Path(output_dir) / f"page-{page_number}.jpg"
                image.write_bytes(b"jpg")
                paths.append(image)
            rendered_ranges.append((first, last))
            return paths

        def upload_page(key, *_args):
            if key.endswith("page-0001.jpg"):
                self.assertNotIn((3, 4), rendered_ranges)
                second_range_release.set()
            return "etag"

        processor.render_pdf_pages = mock.Mock(side_effect=render_range)
        processor.r2.upload.side_effect = upload_page

        with tempfile.TemporaryDirectory() as tmp:
            with mock.patch("worker.worker.PdfReader") as reader_cls:
                reader = reader_cls.return_value
                reader.is_encrypted = False
                reader.pages = [mock.Mock(), mock.Mock(), mock.Mock(), mock.Mock()]
                result = processor.enrich_pdf_job(
                    {"id": "job-2", "document_file_id": "doc-1", "input": {}},
                    Path(tmp),
                )

        self.assertEqual(result["pages_uploaded"], 4)
        self.assertEqual(processor.db.insert_pages.call_count, 4)
        self.assertCountEqual(rendered_ranges, [(1, 2), (3, 4)])

    def test_enrich_removes_uploaded_page_when_manifest_publish_fails(self):
        processor = self._enrich_processor(embeddings_enabled=False)
        processor.db.insert_pages.side_effect = RuntimeError("document deleted")

        with tempfile.TemporaryDirectory() as tmp:
            image = Path(tmp) / "page-1.jpg"
            image.write_bytes(b"jpg")
            processor.render_pdf_pages = mock.Mock(return_value=[image])
            with mock.patch("worker.worker.PdfReader") as reader_cls:
                reader = reader_cls.return_value
                reader.is_encrypted = False
                reader.pages = [mock.Mock()]
                with self.assertRaisesRegex(RuntimeError, "document deleted"):
                    processor.enrich_pdf_job(
                        {"id": "job-2", "document_file_id": "doc-1", "input": {}},
                        Path(tmp),
                    )

        processor.r2.delete.assert_called_once_with(
            "users/user-1/documents/doc-1/pages/page-0001.jpg"
        )

    def test_enrich_incomplete_embeddings_skip_enriched_at(self):
        processor = self._enrich_processor(embed_return=[[0.1] * 8, None])

        with tempfile.TemporaryDirectory() as tmp:
            images = []
            for i in range(1, 3):
                image = Path(tmp) / f"page-{i}.jpg"
                image.write_bytes(b"jpg")
                images.append(image)
            processor.render_pdf_pages = mock.Mock(return_value=images)
            with mock.patch("worker.worker.PdfReader") as reader_cls:
                reader = reader_cls.return_value
                reader.is_encrypted = False
                reader.pages = [mock.Mock(), mock.Mock()]
                result = processor.enrich_pdf_job(
                    {"id": "job-2", "document_file_id": "doc-1", "input": {}},
                    Path(tmp),
                )

        patch = result["_document_patch"]
        self.assertNotIn("enriched_at", patch)
        self.assertEqual(patch["stage_errors"]["enrichment"]["code"], "embedding_incomplete")
        self.assertEqual(result["status"], "visual_ready")
        self.assertFalse(result["enriched"])

    def test_enrich_jina_disabled_preserves_visual_without_enriched(self):
        processor = self._enrich_processor(embeddings_enabled=False)

        with tempfile.TemporaryDirectory() as tmp:
            image = Path(tmp) / "page-1.jpg"
            image.write_bytes(b"jpg")
            processor.render_pdf_pages = mock.Mock(return_value=[image])
            with mock.patch("worker.worker.PdfReader") as reader_cls:
                reader = reader_cls.return_value
                reader.is_encrypted = False
                reader.pages = [mock.Mock()]
                result = processor.enrich_pdf_job(
                    {"id": "job-2", "document_file_id": "doc-1", "input": {}},
                    Path(tmp),
                )

        patch = result["_document_patch"]
        self.assertIn("visual_ready_at", patch)
        self.assertNotIn("enriched_at", patch)
        self.assertEqual(patch["stage_errors"]["enrichment"]["code"], "embedding_unavailable")
        self.assertEqual(result["status"], "visual_ready")
        processor.embeddings.embed_images.assert_not_called()

    def test_enrich_lost_lease_on_null_visual_publish(self):
        processor = self._enrich_processor()
        processor.db.publish_document_visual_ready.return_value = None

        with tempfile.TemporaryDirectory() as tmp:
            image = Path(tmp) / "page-1.jpg"
            image.write_bytes(b"jpg")
            processor.render_pdf_pages = mock.Mock(return_value=[image])
            with mock.patch("worker.worker.PdfReader") as reader_cls:
                reader = reader_cls.return_value
                reader.is_encrypted = False
                reader.pages = [mock.Mock()]
                with self.assertRaises(w.LeaseLostError):
                    processor.enrich_pdf_job(
                        {"id": "job-2", "document_file_id": "doc-1", "input": {}},
                        Path(tmp),
                    )

        processor.embeddings.embed_images.assert_not_called()


class InsertPagesConflictTest(unittest.TestCase):
    def test_insert_pages_supports_ignore_duplicates(self):
        db = w.Supabase.__new__(w.Supabase)
        db.request = mock.Mock(return_value=None)
        db.insert_pages([{"document_file_id": "d", "page_number": 1}], on_conflict="ignore")
        kwargs = db.request.call_args.kwargs
        prefer = kwargs["prefer"]
        self.assertIn("ignore-duplicates", prefer)
        self.assertEqual(kwargs["params"], {"on_conflict": "document_file_id,page_number"})

    def test_insert_pages_defaults_to_merge(self):
        db = w.Supabase.__new__(w.Supabase)
        db.request = mock.Mock(return_value=None)
        db.insert_pages([{"document_file_id": "d", "page_number": 1}])
        kwargs = db.request.call_args.kwargs
        prefer = kwargs["prefer"]
        self.assertIn("merge-duplicates", prefer)
        self.assertEqual(kwargs["params"], {"on_conflict": "document_file_id,page_number"})

    def test_render_page_uses_conflict_ignore(self):
        processor = w.Processor.__new__(w.Processor)
        processor.default_limits = {"max_pdf_pages": 100}
        processor.visual_page_dpi = 144
        processor.pdf_render_workers = 1
        processor.db = mock.Mock()
        processor.db.get_document_file.return_value = {
            "id": "doc-1",
            "user_id": "user-1",
            "attachment_id": "att-1",
            "kind": "pdf",
        }
        processor.db.get_attachment.return_value = {
            "id": "att-1",
            "file_name": "a.pdf",
            "object_key": "raw/a.pdf",
            "content_type": "application/pdf",
            "size_bytes": 1,
            "etag": "e",
        }
        processor.r2 = mock.Mock()
        processor.assert_job_active = mock.Mock()
        processor.estimated_page_pixels = mock.Mock(return_value=(10, 20))
        processor.limit = w.Processor.limit.__get__(processor, w.Processor)

        with tempfile.TemporaryDirectory() as tmp:
            image = Path(tmp) / "page-1.jpg"
            image.write_bytes(b"jpg")
            processor.render_pdf_pages = mock.Mock(return_value=[image])
            with mock.patch("worker.worker.PdfReader") as reader_cls:
                reader = reader_cls.return_value
                reader.is_encrypted = False
                reader.pages = [mock.Mock()]
                processor.render_page_job(
                    {"id": "job-r", "document_file_id": "doc-1", "input": {"page_number": 1}},
                    Path(tmp),
                )

        kwargs = processor.db.insert_pages.call_args.kwargs
        self.assertNotIn("on_conflict", kwargs)

    def test_render_page_converts_office_source_to_temporary_pdf(self):
        processor = w.Processor.__new__(w.Processor)
        processor.default_limits = {"max_pdf_pages": 100}
        processor.visual_page_dpi = 144
        processor.pdf_render_workers = 1
        processor.db = mock.Mock()
        processor.db.get_document_file.return_value = {
            "id": "doc-1",
            "user_id": "user-1",
            "attachment_id": "att-1",
            "kind": "docx",
        }
        processor.db.get_attachment.return_value = {
            "id": "att-1",
            "file_name": "report.docx",
            "object_key": "raw/report.docx",
            "content_type": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "size_bytes": 1,
            "etag": "e",
        }
        processor.r2 = mock.Mock()
        processor.assert_job_active = mock.Mock()
        processor.estimated_page_pixels = mock.Mock(return_value=(10, 20))
        processor.limit = w.Processor.limit.__get__(processor, w.Processor)

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            temporary_pdf = tmp_path / "out" / "report.pdf"
            temporary_pdf.parent.mkdir()
            temporary_pdf.write_bytes(b"pdf")
            image = tmp_path / "page-1.jpg"
            image.write_bytes(b"jpg")
            processor.libreoffice_convert = mock.Mock(return_value=temporary_pdf)
            processor.render_pdf_pages = mock.Mock(return_value=[image])
            with mock.patch("worker.worker.PdfReader") as reader_cls:
                reader = reader_cls.return_value
                reader.is_encrypted = False
                reader.pages = [mock.Mock()]
                processor.render_page_job(
                    {"id": "job-r", "document_file_id": "doc-1", "input": {"page_number": 1}},
                    tmp_path,
                )

        processor.libreoffice_convert.assert_called_once_with(tmp_path / "report.docx", tmp_path, "pdf")
        self.assertEqual(processor.render_pdf_pages.call_args.args[0], temporary_pdf)


class PublishVisualReadyRpcTest(unittest.TestCase):
    def test_publish_document_visual_ready_rpc_payload(self):
        db = w.Supabase.__new__(w.Supabase)
        db.rpc = mock.Mock(return_value={"id": "doc-1"})
        result = db.publish_document_visual_ready(
            "job-1",
            "worker-a",
            "doc-1",
            4,
            {"stage": "visual_ready"},
        )
        self.assertEqual(result["id"], "doc-1")
        db.rpc.assert_called_once_with("klui_publish_document_visual_ready", {
            "p_job_id": "job-1",
            "p_worker_id": "worker-a",
            "p_document_file_id": "doc-1",
            "p_page_count": 4,
            "p_metadata": {"stage": "visual_ready"},
        })


class PdftoppmCommandTest(unittest.TestCase):
    def test_page_range_flags(self):
        cmd = w.build_pdftoppm_command("/tmp/doc.pdf", "/tmp/out/page", 144, first=3, last=3)
        self.assertEqual(cmd[0], "pdftoppm")
        self.assertIn("-f", cmd)
        self.assertEqual(cmd[cmd.index("-f") + 1], "3")
        self.assertEqual(cmd[cmd.index("-l") + 1], "3")
        self.assertEqual(cmd[-2:], ["/tmp/doc.pdf", "/tmp/out/page"])

    def test_render_pdf_pages_uses_requested_range(self):
        processor = w.Processor.__new__(w.Processor)
        processor.visual_page_dpi = 144
        processor.pdf_render_workers = 2
        captured = {}

        def fake_run(cmd, check=True, stdout=None, stderr=None):
            captured["cmd"] = list(cmd)
            return mock.Mock()

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            (out / "page-7.jpg").write_bytes(b"x")
            with mock.patch("worker.worker.subprocess.run", side_effect=fake_run):
                paths = processor.render_pdf_pages(
                    "/tmp/doc.pdf",
                    out,
                    dpi=144,
                    page_count=1,
                    first_page=7,
                    last_page=7,
                )
        self.assertEqual([p.name for p in paths], ["page-7.jpg"])
        self.assertEqual(captured["cmd"][captured["cmd"].index("-f") + 1], "7")
        self.assertEqual(captured["cmd"][captured["cmd"].index("-l") + 1], "7")


class DispatchRoutingTest(unittest.TestCase):
    def test_pdf_creation_uses_docx_renderer_then_libreoffice(self):
        processor = w.Processor.__new__(w.Processor)
        processor.create_js_artifact = mock.Mock()
        processor.create_docx = mock.Mock()
        processor.libreoffice_convert = mock.Mock()
        processor.store_generated = mock.Mock(return_value={"document_file_id": "doc-1"})
        job = {
            "id": "job-1",
            "user_id": "user-1",
            "conversation_id": "conversation-1",
            "job_type": "document.create.pdf",
            "input": {
                "format": "pdf",
                "title": "Report",
                "editor_markdown": "# Report\n\nEditable body.",
            },
        }

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            docx_path = tmp_path / "Report.docx"
            pdf_path = tmp_path / "out" / "Report.pdf"
            processor.create_js_artifact.return_value = docx_path
            processor.libreoffice_convert.return_value = pdf_path

            result = processor.create_job(job, tmp_path)

        self.assertEqual(result, {"document_file_id": "doc-1"})
        processor.create_js_artifact.assert_called_once_with(
            tmp_path, "Report", job["input"], "docx"
        )
        processor.create_docx.assert_not_called()
        processor.libreoffice_convert.assert_called_once_with(docx_path, tmp_path, "pdf")
        processor.store_generated.assert_called_once_with(
            job, tmp_path, pdf_path, "pdf", "application/pdf", "generated", None
        )

    def test_store_generated_keeps_editable_markdown_for_prose_documents(self):
        processor = w.Processor.__new__(w.Processor)
        processor.r2 = mock.Mock()
        processor.r2.upload.return_value = "etag-1"
        processor.db = mock.Mock()
        processor.db.create_attachment.return_value = {"id": "att-1", "file_name": "report.docx"}
        processor.db.create_document_file.return_value = {"id": "doc-1"}
        processor.extract = mock.Mock(return_value=([], {"word_count": 2}))
        processor.object_key = mock.Mock(return_value="users/u/report.docx")
        job = {
            "id": "job-1",
            "user_id": "user-1",
            "conversation_id": "conversation-1",
            "input": {"editor_markdown": "# Report\n\nEditable body."},
        }

        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "report.docx"
            output.write_bytes(b"docx")
            processor.store_generated(
                job, Path(tmp), output, "docx",
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                "generated", None,
            )

        created = processor.db.create_document_file.call_args.args[0]
        updated = processor.db.update_document_file.call_args.args[1]
        self.assertTrue(created["metadata"]["editable"])
        self.assertEqual(updated["metadata"]["editor_markdown"], "# Report\n\nEditable body.")
        self.assertEqual(updated["metadata"]["editor_revision"], 1)

    def test_dispatch_routes_rev3_job_types(self):
        processor = w.Processor.__new__(w.Processor)
        processor.extract_job = mock.Mock(return_value={"ok": "extract"})
        processor.enrich_pdf_job = mock.Mock(return_value={"ok": "enrich"})
        processor.render_page_job = mock.Mock(return_value={"ok": "render"})
        processor.create_job = mock.Mock(return_value={"ok": "create"})
        tmp = Path("/tmp")

        self.assertEqual(
            processor.dispatch({"job_type": "document.extract.pdf"}, tmp)["ok"],
            "extract",
        )
        self.assertEqual(
            processor.dispatch({"job_type": "document.enrich.pdf"}, tmp)["ok"],
            "enrich",
        )
        self.assertEqual(
            processor.dispatch({"job_type": "document.render_page"}, tmp)["ok"],
            "render",
        )
        self.assertEqual(
            processor.dispatch({"job_type": "document.create.docx"}, tmp)["ok"],
            "create",
        )

    def test_handle_job_uses_complete_rpc_and_strips_document_patch(self):
        processor = w.Processor.__new__(w.Processor)
        processor.worker_id = "worker-a"
        processor.lease_seconds = 120
        processor.heartbeat_seconds = 30
        processor.db = mock.Mock()
        processor.db.complete_document_job.return_value = {"job": {"id": "job-1"}}
        processor._lease_heartbeat_loop = mock.Mock()
        processor.assert_job_active = mock.Mock()
        processor.dispatch = mock.Mock(return_value={
            "document_file_id": "doc-1",
            "status": "text_ready",
            "_document_patch": {"text_ready_at": "2026-07-11T00:00:00+00:00", "error": None},
        })

        with mock.patch("worker.worker.tempfile.mkdtemp", return_value=tempfile.mkdtemp()):
            with mock.patch("worker.worker.shutil.rmtree"):
                with mock.patch("worker.worker.threading.Thread") as thread_cls:
                    thread = mock.Mock()
                    thread_cls.return_value = thread
                    processor.handle_job({"id": "job-1", "document_file_id": "doc-1"})

        args = processor.db.complete_document_job.call_args.args
        self.assertEqual(args[0], "job-1")
        self.assertEqual(args[1], "worker-a")
        self.assertEqual(args[2]["status"], "text_ready")
        self.assertNotIn("_document_patch", args[2])
        self.assertEqual(args[3]["text_ready_at"], "2026-07-11T00:00:00+00:00")
        processor.db.fail_document_job.assert_not_called()
        processor.db.update_document_file.assert_not_called()

    def test_handle_job_fails_through_fail_rpc_without_direct_document_fail(self):
        processor = w.Processor.__new__(w.Processor)
        processor.worker_id = "worker-a"
        processor.lease_seconds = 120
        processor.heartbeat_seconds = 30
        processor.db = mock.Mock()
        processor.db.fail_document_job.return_value = {"job": {"id": "job-1"}}
        processor._lease_heartbeat_loop = mock.Mock()
        processor.dispatch = mock.Mock(side_effect=RuntimeError("boom"))

        with mock.patch("worker.worker.tempfile.mkdtemp", return_value=tempfile.mkdtemp()):
            with mock.patch("worker.worker.shutil.rmtree"):
                with mock.patch("worker.worker.threading.Thread") as thread_cls:
                    thread_cls.return_value = mock.Mock()
                    processor.handle_job({"id": "job-1", "document_file_id": "doc-1"})

        processor.db.fail_document_job.assert_called_once()
        error = processor.db.fail_document_job.call_args.args[2]
        self.assertEqual(error["message"], "boom")
        processor.db.update_document_file.assert_not_called()


class RenewJobLeaseTest(unittest.TestCase):
    def test_renew_job_lease_filters_running_and_worker(self):
        db = w.Supabase.__new__(w.Supabase)
        db.url = "https://example.test"
        db.key = "key"
        db.max_attempts = 1
        captured = {}

        def fake_request(path, method="GET", params=None, body=None, prefer=None):
            captured.update({
                "path": path,
                "method": method,
                "params": params,
                "body": body,
                "prefer": prefer,
            })
            return [{"id": "job-1", "status": "running"}]

        db.request = fake_request
        row = db.renew_job_lease("job-1", "worker-a", 120)
        self.assertEqual(row["id"], "job-1")
        self.assertEqual(captured["method"], "PATCH")
        self.assertEqual(captured["params"]["id"], "eq.job-1")
        self.assertEqual(captured["params"]["worker_id"], "eq.worker-a")
        self.assertEqual(captured["params"]["status"], "eq.running")
        self.assertTrue(captured["params"]["lease_until"].startswith("gte."))
        self.assertIn("lease_until", captured["body"])
        self.assertNotIn("status", captured["body"])
        self.assertNotIn("output", captured["body"])

    def test_processor_stops_when_heartbeat_loses_ownership(self):
        processor = w.Processor.__new__(w.Processor)
        processor._lease_lost = threading.Event()
        processor._lease_lost.set()
        with self.assertRaises(w.LeaseLostError):
            processor.assert_job_lease()

    def test_assert_job_active_honors_cancel_requested(self):
        processor = w.Processor.__new__(w.Processor)
        processor.worker_id = "worker-a"
        processor._lease_lost = threading.Event()
        processor.db = mock.Mock()
        processor.db.get_job.return_value = {
            "id": "job-1",
            "status": "running",
            "worker_id": "worker-a",
            "cancel_requested": True,
        }
        with self.assertRaises(w.JobCancelledError):
            processor.assert_job_active("job-1")

    def test_heartbeat_stops_before_an_unrenewed_lease_expires(self):
        processor = w.Processor.__new__(w.Processor)
        processor.worker_id = "worker-a"
        processor.lease_seconds = 120
        processor.heartbeat_seconds = 30
        processor.db = mock.Mock()
        processor.db.renew_job_lease.side_effect = RuntimeError("database unavailable")
        stop_event = mock.Mock()
        stop_event.wait.return_value = False
        lost_event = threading.Event()

        with mock.patch("worker.worker.time.monotonic", side_effect=[0, 91]):
            processor._lease_heartbeat_loop("job-1", stop_event, lost_event)

        self.assertTrue(lost_event.is_set())


class XlsxReadEditTest(unittest.TestCase):
    def test_extract_xlsx_reads_past_row_200_and_emits_bounded_ranges(self):
        processor = w.Processor.__new__(w.Processor)
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "large.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Data"
            sheet.append(["Name", "Value"])
            sheet.append(["Formula", "=SUM(B3:B4)"])
            for row_number in range(3, 261):
                sheet.append([f"row-{row_number}-" + ("x" * 40), row_number])
            workbook.calculation.fullCalcOnLoad = True
            workbook.save(source)
            workbook.close()

            chunks, metadata = processor.extract_xlsx(
                source,
                "user-1",
                "doc-1",
                {"max_xlsx_sheets": 25, "max_xlsx_cells": 250000},
            )

        self.assertGreater(len(chunks), 1)
        self.assertTrue(all(chunk["source_type"] == "sheet_range" for chunk in chunks))
        self.assertTrue(any(260 in chunk["metadata"]["row_numbers"] for chunk in chunks))
        self.assertTrue(all(len(chunk["text"]) < 6000 for chunk in chunks))
        self.assertIn("=SUM(B3:B4)", "\n".join(chunk["text"] for chunk in chunks))
        self.assertNotIn("=> 0", "\n".join(chunk["text"] for chunk in chunks))
        self.assertEqual(metadata["extractor"], "openpyxl_ranges_v2")
        self.assertFalse(metadata["formula_cache_trusted"])
        self.assertGreater(metadata["word_count"], 0)

    def test_edit_xlsx_uses_explicit_operations_and_rejects_unknown_sheets(self):
        processor = w.Processor.__new__(w.Processor)
        with tempfile.TemporaryDirectory() as tmp:
            directory = Path(tmp)
            source = directory / "book.xlsx"
            workbook = Workbook()
            workbook.active.title = "Budget"
            workbook.save(source)
            workbook.close()

            output = processor.edit_xlsx(source, directory, {
                "operations": [
                    {"type": "set_range", "sheet": "Budget", "range": "A1:B2", "values": [["Item", "Cost"], ["Hosting", 20]]},
                    {"type": "set_formula", "sheet": "Budget", "cell": "B3", "formula": "=SUM(B2:B2)"},
                    {"type": "set_number_format", "sheet": "Budget", "range": "B2:B3", "format": "$#,##0.00"},
                ]
            })
            edited = load_workbook(output, data_only=False)
            self.assertEqual(edited["Budget"]["A2"].value, "Hosting")
            self.assertIn(edited["Budget"]["B3"].value, {"=SUM(B2:B2)", "=SUM(B2)"})
            self.assertIn(edited["Budget"]["B2"].number_format, {"$#,##0.00", "\\$#,##0.00"})
            self.assertTrue(edited.calculation.fullCalcOnLoad)
            edited.close()

            with self.assertRaisesRegex(RuntimeError, "xlsx_sheet_not_found"):
                processor.edit_xlsx(source, directory, {
                    "operations": [{"type": "set_cell", "sheet": "Missing", "cell": "A1", "value": "wrong"}]
                })
            original = load_workbook(source)
            self.assertIsNone(original["Budget"]["A1"].value)
            original.close()


class SupabaseRetrySafetyTest(unittest.TestCase):
    def test_non_idempotent_posts_are_not_retried(self):
        db = w.Supabase.__new__(w.Supabase)
        db.url = "https://example.test"
        db.key = "key"
        db.max_attempts = 4
        response = mock.Mock(ok=True, status_code=201, text='[{"id":"row-1"}]')
        response.json.return_value = [{"id": "row-1"}]

        with mock.patch("worker.worker.request_with_retries", return_value=response) as request_mock:
            db.request("attachments", method="POST", body={"id": "row-1"})

        self.assertEqual(request_mock.call_args.kwargs["max_attempts"], 1)

    def test_idempotent_reads_use_configured_retries(self):
        db = w.Supabase.__new__(w.Supabase)
        db.url = "https://example.test"
        db.key = "key"
        db.max_attempts = 4
        response = mock.Mock(ok=True, status_code=200, text="[]")
        response.json.return_value = []

        with mock.patch("worker.worker.request_with_retries", return_value=response) as request_mock:
            db.request("document_files")

        self.assertEqual(request_mock.call_args.kwargs["max_attempts"], 4)


class R2UploadEtagTest(unittest.TestCase):
    def test_upload_uses_put_object_etag_without_head(self):
        r2 = w.R2.__new__(w.R2)
        r2.bucket = "bucket"
        client = mock.Mock()
        client.put_object.return_value = {"ETag": '"abc123"'}
        client.head_object.side_effect = AssertionError("HEAD should not be called")
        r2.client = client

        with tempfile.NamedTemporaryFile(delete=False) as handle:
            handle.write(b"hello")
            path = handle.name
        try:
            etag = r2.upload("key/path.bin", path, "application/octet-stream")
        finally:
            Path(path).unlink(missing_ok=True)

        self.assertEqual(etag, "abc123")
        client.put_object.assert_called_once()
        client.head_object.assert_not_called()
        client.upload_file.assert_not_called()

    def test_delete_removes_the_exact_r2_object(self):
        r2 = w.R2.__new__(w.R2)
        r2.bucket = "bucket"
        r2.client = mock.Mock()

        r2.delete("users/user-1/page.jpg")

        r2.client.delete_object.assert_called_once_with(
            Bucket="bucket",
            Key="users/user-1/page.jpg",
        )


class JinaBatchOrderingTest(unittest.TestCase):
    def test_embed_images_preserves_order_with_parallel_batches(self):
        embeddings = w.JinaEmbeddings.__new__(w.JinaEmbeddings)
        embeddings.api_key = "test"
        embeddings.model = "model"
        embeddings.dimensions = 768
        embeddings.endpoint = "https://example.test/embeddings"
        embeddings.batch_size = 2
        embeddings.batch_concurrency = 2
        embeddings.max_attempts = 1

        def embed_batch(batch):
            return [(original_index, f"emb-{original_index}") for original_index, _ in batch]

        embeddings._embed_batch = embed_batch

        with tempfile.TemporaryDirectory() as tmp:
            paths = []
            for index in range(5):
                path = Path(tmp) / f"page-{index}.jpg"
                path.write_bytes(b"x" * 10)
                paths.append(path)
            result = embeddings.embed_images(paths, batch_size=2, batch_concurrency=2)

        self.assertEqual(result, ["emb-0", "emb-1", "emb-2", "emb-3", "emb-4"])


class HealthcheckTest(unittest.TestCase):
    def test_healthcheck_requires_edgeparse(self):
        from worker import healthcheck

        with mock.patch("worker.healthcheck.importlib.util.find_spec", return_value=None):
            self.assertEqual(healthcheck.main(), 1)

    def test_healthcheck_requires_pdftoppm(self):
        from worker import healthcheck

        def fake_which(name):
            if name == "pdftoppm":
                return None
            return "/usr/bin/" + name

        with mock.patch("worker.healthcheck.importlib.util.find_spec", return_value=object()), \
             mock.patch("worker.healthcheck.shutil.which", side_effect=fake_which):
            self.assertEqual(healthcheck.main(), 1)

        with mock.patch("worker.healthcheck.importlib.util.find_spec", return_value=object()), \
             mock.patch("worker.healthcheck.shutil.which", side_effect=lambda name: "/bin/" + name):
            self.assertEqual(healthcheck.main(), 0)


if __name__ == "__main__":
    unittest.main()
