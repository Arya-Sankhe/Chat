import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from worker.xlsx_generator import create_xlsx_workbook, sanitize_sheet_name, validate_formula


class XlsxGeneratorTest(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.path = Path(self.tmp.name) / "out.xlsx"

    def test_formula_row_correctness_header_is_row_1(self):
        create_xlsx_workbook(self.path, {
            "title": "Sales",
            "data": {
                "sheets": [{
                    "name": "Sales",
                    "rows": [
                        ["Region", "Revenue"],
                        ["North", 100],
                        ["South", 200],
                        ["Total", "=SUM(B2:B3)"],
                    ],
                    "columns": [{}, {"format": "number"}],
                }]
            },
        })
        ws = load_workbook(self.path)["Sales"]
        self.assertEqual(ws["A1"].value, "Region")
        self.assertEqual(ws["B2"].value, 100)
        self.assertEqual(ws["B4"].value, "=SUM(B2:B3)")
        self.assertEqual(ws["B4"].number_format, "#,##0.00")

    def test_leading_zero_text_preserved(self):
        create_xlsx_workbook(self.path, {
            "data": {
                "sheets": [{
                    "name": "Codes",
                    "rows": [["SKU", "Qty"], ["00123", 2]],
                    "columns": [{"format": "text"}, {"format": "integer"}],
                }]
            }
        })
        ws = load_workbook(self.path)["Codes"]
        self.assertEqual(ws["A2"].value, "00123")
        self.assertEqual(ws["B2"].value, 2)

    def test_percent_is_decimal_fraction(self):
        create_xlsx_workbook(self.path, {
            "data": {
                "sheets": [{
                    "name": "Rates",
                    "rows": [["Label", "Rate"], ["Win", 0.45], ["Loss", "12.5%"]],
                    "columns": [{}, {"format": "percent"}],
                }]
            }
        })
        ws = load_workbook(self.path)["Rates"]
        self.assertEqual(ws["B2"].value, 0.45)
        self.assertEqual(ws["B2"].number_format, "0.0%")
        self.assertEqual(ws["B3"].value, 0.125)

    def test_cover_only_when_explicit(self):
        no_cover = Path(self.tmp.name) / "no-cover.xlsx"
        create_xlsx_workbook(no_cover, {
            "title": "Multi",
            "data": {
                "sheets": [
                    {"name": "A", "rows": [["H"], ["1"]]},
                    {"name": "B", "rows": [["H"], ["2"]]},
                ]
            },
        })
        self.assertEqual(load_workbook(no_cover).sheetnames, ["A", "B"])

        with_cover = Path(self.tmp.name) / "with-cover.xlsx"
        create_xlsx_workbook(with_cover, {
            "title": "Multi",
            "data": {
                "cover": {"subtitle": "Overview", "metrics": [{"label": "Total", "value": 3}]},
                "sheets": [
                    {"name": "A", "description": "First", "rows": [["H"], ["1"]]},
                    {"name": "B", "rows": [["H"], ["2"]]},
                ],
            },
        })
        wb = load_workbook(with_cover)
        self.assertEqual(wb.sheetnames, ["Cover", "A", "B"])
        cover = wb["Cover"]
        self.assertEqual(cover["B2"].value, "Multi")
        values = [cell for row in cover.iter_rows(values_only=True) for cell in row]
        self.assertIn("Total", values)
        self.assertIn("First", values)

    def test_empty_payload_is_rejected_instead_of_creating_fake_summary(self):
        with self.assertRaisesRegex(ValueError, "xlsx_requires_sheet_data"):
            create_xlsx_workbook(self.path, {
                "title": "Promised workbook",
                "instructions": "Create pricing and scenario tables",
                "data": {"sheets": [{"name": "Summary", "rows": []}]},
            })
        self.assertFalse(self.path.exists())

    def test_unique_names_including_cover_collision(self):
        create_xlsx_workbook(self.path, {
            "title": "Clash",
            "data": {
                "cover": {"subtitle": "Index"},
                "sheets": [
                    {"name": "Cover", "rows": [["H"], ["1"]]},
                    {"name": "cover", "rows": [["H"], ["1"]]},
                    {"name": "Q1: Revenue / Costs", "rows": [["H"], ["2"]]},
                    {"name": "Q1: Revenue / Costs", "rows": [["H"], ["3"]]},
                ],
            },
        })
        names = load_workbook(self.path).sheetnames
        self.assertEqual(names[0], "Cover")
        self.assertEqual(names[1], "Cover-2")
        self.assertEqual(names[2], "cover-3")
        self.assertEqual(names[3], sanitize_sheet_name("Q1: Revenue / Costs"))
        self.assertEqual(names[4], f"{sanitize_sheet_name('Q1: Revenue / Costs')}-2")
        self.assertEqual(len(names), len(set(names)))

    def test_charts_and_conditional_formats(self):
        create_xlsx_workbook(self.path, {
            "theme": "business",
            "data": {
                "sheets": [{
                    "name": "Sales",
                    "rows": [
                        ["Region", "Revenue", "Cost"],
                        ["North", 120, 80],
                        ["South", 95, 60],
                    ],
                    "columns": [{}, {"format": "currency", "symbol": "$"}, {"format": "number"}],
                    "charts": [{
                        "type": "bar",
                        "title": "Revenue by Region",
                        "categories_column": 1,
                        "series_columns": [2, 3],
                        "x_axis_title": "Region",
                        "y_axis_title": "Amount",
                    }],
                    "conditional_formats": [{"column": 2, "type": "data_bar"}],
                }]
            },
        })
        ws = load_workbook(self.path)["Sales"]
        self.assertEqual(len(ws._charts), 1)
        self.assertTrue(ws.freeze_panes)
        self.assertEqual(len(ws.tables), 1)
        self.assertTrue(next(iter(ws.tables.values())).autoFilter)
        self.assertGreaterEqual(len(ws.conditional_formatting._cf_rules), 1)

    def test_invalid_formula_becomes_text_without_failing_workbook(self):
        create_xlsx_workbook(self.path, {
            "data": {
                "sheets": [{
                    "name": "Bad",
                    "rows": [["Formula", "Value"], ["=FILTER(A2:A9, A2:A9>0)", 42]],
                }]
            }
        })
        ws = load_workbook(self.path)["Bad"]
        self.assertEqual(ws["A2"].value, "=FILTER(A2:A9, A2:A9>0)")
        self.assertEqual(ws["A2"].data_type, "s")
        self.assertEqual(ws["B2"].value, 42)
        with self.assertRaises(ValueError):
            validate_formula("=WEBSERVICE(\"http://evil\")")
        with self.assertRaises(ValueError):
            validate_formula("=XLOOKUP(A1,B:B,C:C)")
        with self.assertRaises(ValueError):
            validate_formula("=#REF!+1")
        self.assertEqual(validate_formula("=SUM(B2:B3)"), "=SUM(B2:B3)")
        self.assertEqual(validate_formula("=IFERROR(A2/B2,0)"), "=IFERROR(A2/B2,0)")
        self.assertEqual(validate_formula('=COUNTIF(A2:A10,"https://example.com")'), '=COUNTIF(A2:A10,"https://example.com")')
        self.assertEqual(validate_formula('=IFERROR(A1,"#N/A")'), '=IFERROR(A1,"#N/A")')

    def test_cross_sheet_formulas_follow_sanitized_names(self):
        create_xlsx_workbook(self.path, {
            "data": {"sheets": [
                {"name": "Q1: Revenue (Actual)", "rows": [["Value"], [10]]},
                {"name": "Summary", "rows": [["Total"], ["='q1: revenue (actual)'!A2"]]},
            ]},
        })
        formula = load_workbook(self.path)["Summary"]["A2"].value
        self.assertEqual(formula, "='Q1-Revenue-Actual'!A2")

    def test_sheet_rewrite_does_not_match_inside_a_longer_name(self):
        create_xlsx_workbook(self.path, {
            "data": {"sheets": [
                {"name": "Sales Raw", "rows": [["Value"], [10]]},
                {"name": "MySales Raw", "rows": [["Value"], [20]]},
                {"name": "Summary", "rows": [["Total"], ["='MySales Raw'!A2"]]},
            ]},
        })
        formula = load_workbook(self.path)["Summary"]["A2"].value
        self.assertEqual(formula, "='MySales-Raw'!A2")

    def test_chart_excludes_trailing_total_row(self):
        create_xlsx_workbook(self.path, {
            "data": {"sheets": [{
                "name": "Sales",
                "rows": [["Region", "Revenue"], ["North", 10], ["South", 20], ["Total", "=SUM(B2:B3)"]],
                "charts": [{"type": "bar", "categories_column": 1, "series_columns": [2]}],
            }]},
        })
        chart = load_workbook(self.path)["Sales"]._charts[0]
        self.assertEqual(chart.series[0].cat.strRef.f, "Sales!$A$2:$A$3")
        self.assertEqual(chart.series[0].val.numRef.f, "Sales!$B$2:$B$3")

    def test_legacy_rows_and_tables(self):
        legacy_rows = Path(self.tmp.name) / "legacy-rows.xlsx"
        create_xlsx_workbook(legacy_rows, {
            "title": "Legacy",
            "data": {"rows": [["Name", "Value"], ["A", 1]]},
        })
        self.assertEqual(load_workbook(legacy_rows).sheetnames, ["Summary"])
        self.assertEqual(load_workbook(legacy_rows)["Summary"]["A1"].value, "Name")

        top_rows = Path(self.tmp.name) / "top-rows.xlsx"
        create_xlsx_workbook(top_rows, {"rows": [["X", "Y"], ["1", "2"]]})
        self.assertEqual(load_workbook(top_rows)["Summary"]["B1"].value, "Y")

        tables = Path(self.tmp.name) / "tables.xlsx"
        create_xlsx_workbook(tables, {
            "tables": [{"headers": ["Col"], "rows": [["z"]]}],
        })
        self.assertEqual(load_workbook(tables)["Summary"]["A2"].value, "z")

    def test_unspecified_columns_preserve_strings(self):
        create_xlsx_workbook(self.path, {
            "data": {
                "sheets": [{
                    "name": "Raw",
                    "rows": [["Code"], ["001"]],
                }]
            }
        })
        self.assertEqual(load_workbook(self.path)["Raw"]["A2"].value, "001")


if __name__ == "__main__":
    unittest.main()
